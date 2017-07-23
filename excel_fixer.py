#!/usr/bin/python3
import openpyxl
import sys
import win32com.client



#TODO: openpyxl remove bold, remove extras, remove newlines, combine lists
def fix_extra(filename):
    "Fixes the given xlsx file"
    book = openpyxl.load_workbook(filename, data_only=True)
    
    # Unbolding style
    # unbold = openpyxl.styles.Font(bold=False)
    
    # Removing extra line from tabstract.io


    # Recording maximum row of the first sheet
    maximumr = book.worksheets[0].max_row + 1

    for t in range(1, len(book.worksheets)):
        for j in range(1, book.worksheets[t].max_row):
            for k in range(1, book.worksheets[t].max_column):
                book.worksheets[0].cell(row=maximumr, column=k).value = book.worksheets[t].cell(row=j, column=k).value
            maximumr = book.worksheets[0].max_row + 1
    book.save(filename)

if len(sys.argv) < 2:
    print('Arguments error')

try:
    for i in range(1, len(sys.argv)):
        fix_extra(sys.argv[i])
except OSError:
    print('Add filename as an argument')

