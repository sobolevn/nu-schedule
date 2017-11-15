#!/usr/bin/python3

import postgresql
from openpyxl import *

book = load_workbook("sst_shss_fall2017.xlsx", data_only=True, read_only=True)
sheet = book.worksheets[0]

db = postgresql.open("pq://postgres:12345@localhost:5432/postgres")

db.execute("CREATE TABLE coords (id SERIAL PRIMARY KEY, "
            "courseabbr CHAR(640), st CHAR(640), coursetitle CHAR(640), crus CHAR(640), crects CHAR(640), startdate CHAR(640), enddate CHAR(640), days CHAR(640), t1me CHAR(640), enr CHAR(640), cap CHAR(640), faculty CHAR(640), room CHAR(640) )")

for i in range(2, sheet.max_row):
    ins = db.prepare("INSERT INTO coords (courseabbr, st, coursetitle, crus, crects, startdate, enddate, days, t1me, enr, cap, faculty, room) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)")
    ins(str(sheet.cell(row=i, column=1).value),
    str(sheet.cell(row=i, column=2).value)
    , str(sheet.cell(row=i, column=3).value)
    , str(sheet.cell(row=i, column=4).value)
    , str(sheet.cell(row=i, column=5).value)
    , str(sheet.cell(row=i, column=6).value)
    , str(sheet.cell(row=i, column=7).value)
    , str(sheet.cell(row=i, column=8).value)
    , str(sheet.cell(row=i, column=9).value)
    , str(sheet.cell(row=i, column=10).value)
    , str(sheet.cell(row=i, column=11).value)
    , str(sheet.cell(row=i, column=12).value)
    , str(sheet.cell(row=i, column=13).value)
    )
    # (db.prepare("INSERT INTO coords(courseabbr) VALUES($1)"))()
    # (db.prepare("INSERT INTO coords(st) VALUES($1)"))()
    # (db.prepare("INSERT INTO coords(coursetitle) VALUES($1)"))(str(sheet.cell(row=i, column=3).value))
    # (db.prepare("INSERT INTO coords(crus) VALUES($1)"))(str(sheet.cell(row=i, column=4).value))
    # (db.prepare("INSERT INTO coords(crects) VALUES($1)"))(str(sheet.cell(row=i, column=5).value))
    # (db.prepare("INSERT INTO coords(startdate) VALUES($1)"))(str(sheet.cell(row=i, column=6).value))
    # (db.prepare("INSERT INTO coords(enddate) VALUES($1)"))(str(sheet.cell(row=i, column=7).value))
    # (db.prepare("INSERT INTO coords(days) VALUES($1)"))(str(sheet.cell(row=i, column=8).value))
    # (db.prepare("INSERT INTO coords(t1me) VALUES($1)"))(str(sheet.cell(row=i, column=9).value))
    # (db.prepare("INSERT INTO coords(enr) VALUES($1)"))(str(sheet.cell(row=i, column=10).value))
    # (db.prepare("INSERT INTO coords(cap) VALUES($1)"))(str(sheet.cell(row=i, column=11).value))
    # (db.prepare("INSERT INTO coords(faculty) VALUES($1)"))(str(sheet.cell(row=i, column=12).value))
    # (db.prepare("INSERT INTO coords(room) VALUES($1)"))(str(sheet.cell(row=i, column=13).value))

    